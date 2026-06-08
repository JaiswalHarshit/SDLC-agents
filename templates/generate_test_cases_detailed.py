"""
TEMPLATE: Detailed Test Cases Excel Generator (one sheet per TC with full steps)
=================================================================================
Usage by test-plan-writer agent:
  1. Populate the TEST_CASES list below with actual test case data.
  2. Set OUTPUT_PATH to the desired output file path.
  3. Run: python3 generate_test_cases_detailed.py

Each test case dict must have ALL of the following keys:
  id                    – e.g. "TC-NF-01"
  title                 – full one-line title
  priority              – "P1" | "P2" | "P3"
  category              – "New Feature" | "Changed Behaviour" | "Negative" | "Regression"
  test_area             – functional area (e.g. "Staffing", "Contact Manager")
  test_type             – "End-to-End" | "UI Flow" | "API" | "DB Verification" |
                          "Configuration" | "Permission / Role"
  acceptance_criteria   – which ACs this covers (e.g. "AC-1, AC-3" or "N/A")
  derived_from          – "Jira" | "Design Doc: <page, section>" | "Code diff" | "Inferred"
  login_role            – role / username for test execution (e.g. "Admin / ts_admin")
  objective             – one sentence: what this test proves
  preconditions         – list of strings, one per bullet
  test_data             – list of strings, one per bullet (may include inline SQL)
  sql_setup             – multi-line SQL string for data setup, or None
  json_setup            – multi-line JSON string for request payload / config setup, or None
  steps                 – list of (action, navigation_input, expected_result) tuples
  sql_verify            – multi-line SQL string for post-execution verification, or None
  json_verify           – multi-line JSON string for expected response / config verification, or None
  pass_criteria         – explicit pass statement string
  fail_criteria         – explicit fail statement string
  notes                 – caveats / gotchas string (or "—")
"""

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Configuration — edit these two lines ─────────────────────────────────────
OUTPUT_PATH = "/path/to/output/Test_Cases_Detailed.xlsx"
JIRA_TICKETS = "<JIRA_TICKET(S)>"
PRS          = "PR #<NUMBER> (<description>)"

# ── Colour palette ────────────────────────────────────────────────────────────
CLR = {
    "hdr_bg":      "1F4E79",
    "hdr_fg":      "FFFFFF",
    "p1_bg":       "C00000",
    "p2_bg":       "ED7D31",
    "p3_bg":       "808080",
    "p1_fg":       "FFFFFF",
    "p2_fg":       "FFFFFF",
    "p3_fg":       "FFFFFF",
    "step_hdr":    "2E75B6",
    "step_hdr_fg": "FFFFFF",
    "odd_row":     "DEEAF1",
    "even_row":    "FFFFFF",
    "section_bg":  "BDD7EE",
    "section_fg":  "1F4E79",
    "pass_bg":     "E2EFDA",
    "fail_bg":     "FCE4D6",
    "sql_bg":      "F2F2F2",
    "json_bg":     "FFF8E1",
    "checklist_hdr": "1F4E79",
    "nf_cat":      "D6E4BC",
    "cb_cat":      "D6DCE4",
    "neg_cat":     "FCE4D6",
    "reg_cat":     "FFF2CC",
    "white":       "FFFFFF",
    "grey_lt":     "F2F2F2",
    "blue_pale":   "DEEAF1",
    "result_bg":   "FFFBE6",
    "input_bg":    "F5F5F5",
    "alt_row":     "F5F5F5",
}

CAT_COLOR = {
    "New Feature":        CLR["nf_cat"],
    "Changed Behaviour":  CLR["cb_cat"],
    "Negative":           CLR["neg_cat"],
    "Regression":         CLR["reg_cat"],
}

PRIORITY_COLORS = {
    "P1": CLR["p1_bg"],
    "P2": CLR["p2_bg"],
    "P3": CLR["p3_bg"],
}

# ── Required keys for each test case dict ────────────────────────────────────
REQUIRED_TC_KEYS = {
    "id", "title", "priority", "category", "test_area", "test_type",
    "acceptance_criteria", "derived_from", "login_role",
    "objective", "preconditions", "test_data", "steps",
    "pass_criteria", "fail_criteria",
}

# ── Step table columns ────────────────────────────────────────────────────────
STEP_COLS   = ["Step #", "Action", "Navigation / Input", "Expected Result", "Actual Result", "Pass / Fail"]
STEP_WIDTHS = [7, 28, 32, 32, 20, 10]
STEP_COL_ACTUAL = 5
STEP_COL_PASSFAIL = 6

# ── Checklist column definitions ─────────────────────────────────────────────
CL_HEADERS = ["TC ID", "Title", "Category", "Test Type", "Priority",
              "Tester Name", "Execution Date", "Build / Version",
              "Result (Pass/Fail/Blocked)", "Notes / Defects"]
CL_WIDTHS  = [10, 48, 18, 16, 9, 18, 16, 16, 24, 30]

# Checklist column indices (1-based)
CL_COL = {label: i for i, label in enumerate(CL_HEADERS, 1)}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _side(style="thin", color="000000"):
    return Side(style=style, color=color)


def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, italic=False, name="Calibri", underline=None):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name, underline=underline)


def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _estimate_height(text, col_width=35, line_height=15, min_height=18, max_height=150):
    """Estimate row height based on text content and column width."""
    if not text:
        return min_height
    lines = len(text) // col_width + text.count("\n") + 1
    return min(max(lines * line_height, min_height), max_height)


def validate_test_case(tc, index):
    """Validate that a test case dict has all required keys."""
    missing = REQUIRED_TC_KEYS - set(tc.keys())
    if missing:
        raise ValueError(
            f"Test case at index {index} (id={tc.get('id', '???')}) "
            f"is missing required keys: {sorted(missing)}"
        )


def section_label(ws, row, label, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill(CLR["section_bg"])
    c.font = _font(bold=True, color=CLR["section_fg"])
    c.alignment = _align("left")
    c.border = thin_border()


def kv_row(ws, row, key, value, ncols=6):
    k = ws.cell(row=row, column=1, value=key)
    k.font = _font(bold=True)
    k.fill = fill(CLR["grey_lt"])
    k.border = thin_border()
    k.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    v = ws.cell(row=row, column=2, value=value)
    v.font = _font()
    v.border = thin_border()
    v.alignment = _align("left")


def step_table_header(ws, row):
    for col, label in enumerate(STEP_COLS, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = fill(CLR["step_hdr"])
        c.font = _font(bold=True, color=CLR["step_hdr_fg"])
        c.alignment = _align("center")
        c.border = thin_border()


def step_row(ws, row, step_num, action, nav, expected):
    row_fill = fill(CLR["odd_row"] if step_num % 2 == 1 else CLR["even_row"])
    for col, val in enumerate([step_num, action, nav, expected, "", ""], 1):
        c = ws.cell(row=row, column=col, value=val)
        if col in (STEP_COL_ACTUAL, STEP_COL_PASSFAIL):
            c.fill = fill(CLR["input_bg"])
        else:
            c.fill = row_fill
        c.font = _font()
        c.alignment = _align("left")
        c.border = thin_border()


def sql_block(ws, row, sql_text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=sql_text)
    c.fill = fill(CLR["sql_bg"])
    c.font = _font(color="1F4E79", size=9, name="Courier New")
    c.alignment = _align("left", v="top")
    c.border = thin_border()


def json_block(ws, row, json_text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=json_text)
    c.fill = fill(CLR["json_bg"])
    c.font = _font(color="6D4C00", size=9, name="Courier New")
    c.alignment = _align("left", v="top")
    c.border = thin_border()


def set_step_col_widths(ws):
    for col, w in enumerate(STEP_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Build one TC sheet ────────────────────────────────────────────────────────
def build_tc_sheet(wb, tc):
    ws = wb.create_sheet(title=tc["id"])
    ws.sheet_view.showGridLines = False
    set_step_col_widths(ws)

    # Title bar
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{tc['id']}  |  {tc['title']}"
    t.fill = fill(CLR["hdr_bg"])
    t.font = _font(bold=True, color=CLR["hdr_fg"], size=12)
    t.alignment = _align("left")
    t.border = thin_border()
    ws.row_dimensions[1].height = 24

    row = 2
    meta = [
        ("Priority",              tc["priority"]),
        ("Category",              tc["category"]),
        ("Test Area",             tc["test_area"]),
        ("Test Type",             tc["test_type"]),
        ("Acceptance Criteria",   tc["acceptance_criteria"]),
        ("Derived From",          tc["derived_from"]),
        ("Login Role / User",     tc["login_role"]),
        ("Objective",             tc["objective"]),
        ("Jira Ticket(s)",        JIRA_TICKETS),
        ("PRs",                   PRS),
    ]
    for key, val in meta:
        kv_row(ws, row, key, val)
        ws.row_dimensions[row].height = 30 if key == "Objective" else 18
        row += 1

    # Pre-conditions
    section_label(ws, row, "PRE-CONDITIONS")
    row += 1
    for pc in tc["preconditions"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=f"  • {pc}")
        c.font = _font()
        c.fill = fill(CLR["white"])
        c.border = thin_border()
        c.alignment = _align("left")
        ws.row_dimensions[row].height = 18
        row += 1

    # Test data
    section_label(ws, row, "TEST DATA REQUIRED")
    row += 1
    for td in tc["test_data"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=td)
        c.font = _font()
        c.fill = fill(CLR["white"])
        c.border = thin_border()
        c.alignment = _align("left", v="top")
        ws.row_dimensions[row].height = _estimate_height(td, min_height=18, max_height=120)
        row += 1

    # SQL setup
    if tc.get("sql_setup"):
        section_label(ws, row, "SQL – TEST DATA SETUP / VERIFICATION")
        row += 1
        sql_block(ws, row, tc["sql_setup"])
        ws.row_dimensions[row].height = max(60, 12 * tc["sql_setup"].count("\n"))
        row += 1

    # JSON setup
    if tc.get("json_setup"):
        section_label(ws, row, "JSON – REQUEST PAYLOAD / CONFIG SETUP")
        row += 1
        json_block(ws, row, tc["json_setup"])
        ws.row_dimensions[row].height = max(60, 12 * tc["json_setup"].count("\n"))
        row += 1

    # Steps
    section_label(ws, row, "EXECUTION STEPS")
    row += 1
    step_table_header(ws, row)
    ws.row_dimensions[row].height = 18
    row += 1
    for i, (action, nav, expected) in enumerate(tc["steps"], 1):
        step_row(ws, row, i, action, nav, expected)
        lines = max(
            len(action) // 35 + action.count("\n") + 1,
            len(nav) // 35 + nav.count("\n") + 1,
            len(expected) // 35 + expected.count("\n") + 1,
        )
        ws.row_dimensions[row].height = min(max(lines * 15, 30), 150)
        row += 1

    # SQL verify
    if tc.get("sql_verify"):
        section_label(ws, row, "POST-EXECUTION DB VERIFICATION")
        row += 1
        sql_block(ws, row, tc["sql_verify"])
        ws.row_dimensions[row].height = max(60, 12 * tc["sql_verify"].count("\n"))
        row += 1

    # JSON verify
    if tc.get("json_verify"):
        section_label(ws, row, "JSON – EXPECTED RESPONSE / VERIFICATION")
        row += 1
        json_block(ws, row, tc["json_verify"])
        ws.row_dimensions[row].height = max(60, 12 * tc["json_verify"].count("\n"))
        row += 1

    # Pass / Fail / Notes
    section_label(ws, row, "PASS / FAIL CRITERIA & NOTES")
    row += 1
    for key, val in [("PASS Criteria", tc["pass_criteria"]),
                     ("FAIL Criteria", tc["fail_criteria"]),
                     ("Notes / Gotchas", tc.get("notes", "—"))]:
        kv_row(ws, row, key, val)
        ws.row_dimensions[row].height = _estimate_height(
            val, col_width=100, min_height=30, max_height=120
        )
        row += 1

    ws.freeze_panes = "A2"


# ── Build Checklist sheet ─────────────────────────────────────────────────────
def build_checklist(wb, test_cases):
    ws = wb.create_sheet(title="Checklist", index=0)
    ws.sheet_view.showGridLines = False
    ncols = len(CL_HEADERS)
    today = datetime.date.today().isoformat()

    # Title
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws.cell(row=1, column=1, value=f"TEST EXECUTION CHECKLIST — {JIRA_TICKETS}")
    t.fill = fill(CLR["checklist_hdr"])
    t.font = _font(bold=True, color=CLR["hdr_fg"], size=13)
    t.alignment = _align("center")
    t.border = thin_border()
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    s = ws.cell(row=2, column=1, value=f"{PRS}  |  Date: {today}")
    s.fill = fill(CLR["section_bg"])
    s.font = _font(italic=True, color=CLR["section_fg"])
    s.alignment = _align("center")
    s.border = thin_border()
    ws.row_dimensions[2].height = 16

    # Column headers
    for col, (h, w) in enumerate(zip(CL_HEADERS, CL_WIDTHS), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = fill(CLR["checklist_hdr"])
        c.font = _font(bold=True, color=CLR["hdr_fg"])
        c.alignment = _align("center", wrap=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    dv = DataValidation(
        type="list", formula1='"Pass,Fail,Blocked,Not Run,N/A"',
        allow_blank=True, showDropDown=False,
    )
    ws.add_data_validation(dv)

    for i, tc in enumerate(test_cases):
        row = 4 + i
        cat_fill = fill(CAT_COLOR.get(tc["category"], CLR["white"]))
        pri_color = PRIORITY_COLORS.get(tc["priority"], CLR["p3_bg"])

        values = [tc["id"], tc["title"], tc["category"], tc["test_type"],
                  tc["priority"], "", "", "", "", ""]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border()
            c.alignment = _align("left" if col != CL_COL["Priority"] else "center")
            c.font = _font()

            if col == CL_COL["TC ID"]:
                c.fill = fill(CLR["blue_pale"])
                c.font = _font(bold=True, color=CLR["section_fg"], underline="single")
                c.hyperlink = f"#'{tc['id']}'!A1"
            elif col == CL_COL["Category"]:
                c.fill = cat_fill
            elif col == CL_COL["Priority"]:
                c.fill = fill(pri_color)
                c.font = _font(bold=True, color=CLR["hdr_fg"])
            elif col == CL_COL["Result (Pass/Fail/Blocked)"]:
                c.fill = fill(CLR["result_bg"])
                dv.add(c)
            else:
                c.fill = fill(CLR["white"] if i % 2 == 0 else CLR["alt_row"])
        ws.row_dimensions[row].height = 22

    # Auto-filter on checklist data
    last_data_row = 3 + len(test_cases)
    if test_cases:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{last_data_row}"

    legend_row = 4 + len(test_cases) + 1
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=ncols)
    lg = ws.cell(
        row=legend_row, column=1,
        value="LEGEND:   Green = New Feature   |   Blue-Grey = Changed Behaviour   "
              "|   Orange = Negative   |   Yellow = Regression   "
              "|   P1 = Must pass   |   P2 = Should pass   |   P3 = Nice to verify   "
              "|   Click TC ID to jump to sheet",
    )
    lg.fill = fill(CLR["grey_lt"])
    lg.font = _font(italic=True, size=9, color="595959")
    lg.alignment = _align("left")
    lg.border = thin_border()

    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
#  TEST CASES — populate one dict per test case
# ══════════════════════════════════════════════════════════════════════════════
TEST_CASES = [
    # {
    #     "id":                    "TC-NF-01",
    #     "title":                 "Short descriptive title",
    #     "priority":              "P1",
    #     "category":              "New Feature",
    #     "test_area":             "Staffing",
    #     "test_type":             "End-to-End",
    #     "acceptance_criteria":   "AC-1, AC-2",
    #     "derived_from":          "Jira",
    #     "login_role":            "Admin / ts_admin",
    #     "objective":             "One sentence: what this test proves.",
    #     "preconditions":         ["First condition", "Second condition"],
    #     "test_data":             ["Record name, field value, SQL to create/verify"],
    #     "sql_setup":             "SELECT ...\n-- Expected: ...",   # or None
    #     "json_setup":            '{\n  "field": "value"\n}',      # or None
    #     "steps": [
    #         ("Open browser and log in", "Enter URL -> credentials -> Sign In", "Dashboard loads"),
    #         ("Navigate to ...",          "Click Setup -> ...",                  "Page loads"),
    #     ],
    #     "sql_verify":            "SELECT ...\n-- Expected: ...",   # or None
    #     "json_verify":           '// Expected response:\n{\n  "status": "ok"\n}',  # or None
    #     "pass_criteria":         "Explicit pass statement.",
    #     "fail_criteria":         "Explicit fail statement.",
    #     "notes":                 "Caveats or gotchas.",
    # },
]

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not TEST_CASES:
        print("WARNING: TEST_CASES is empty — no test case sheets will be generated.")

    for i, tc in enumerate(TEST_CASES):
        validate_test_case(tc, i)

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir and not os.path.isdir(out_dir):
        print(f"ERROR: Output directory does not exist: {out_dir}", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)
    build_checklist(wb, TEST_CASES)
    for tc in TEST_CASES:
        build_tc_sheet(wb, tc)

    try:
        wb.save(OUTPUT_PATH)
    except OSError as e:
        print(f"ERROR: Failed to save workbook: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
