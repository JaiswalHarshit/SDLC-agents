"""
TEMPLATE: Test Scenarios Excel Generator (no detailed steps)
============================================================
Usage by test-plan-writer agent:
  1. Populate the SCENARIOS list below with actual test case data.
  2. Set OUTPUT_PATH to the desired output file path.
  3. Set TITLE / SUBTITLE to reflect the Jira tickets and PRs.
  4. Run: python3 generate_test_scenarios.py

Each scenario dict must have ALL of the following keys:
  id            – e.g. "TC-NF-01"
  title         – full one-line title
  priority      – "P1" | "P2" | "P3"
  category      – "New Feature" | "Changed Behaviour" | "Negative" | "Regression"
  test_area     – functional area (e.g. "Staffing", "Contact Manager", "Organization")
  type          – "End-to-End" | "UI Flow" | "API" | "DB Verification" |
                  "Configuration" | "Permission / Role"
  ac_coverage   – acceptance criteria covered (e.g. "AC-1, AC-3" or "N/A")
  objective     – one sentence: what this test proves
  preconditions – numbered string (use \n between items)
  test_data     – string describing required data / SQL snippets
  pass_criteria – explicit pass statement
  fail_criteria – explicit fail statement
  notes         – caveats, boundary conditions, tool tips (or "—")
"""

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Configuration — edit these three lines ───────────────────────────────────
TITLE       = "TEST SCENARIOS — <JIRA_TICKET(S)> — <Feature Name>"
SUBTITLE    = "PRs: #<PR_NUMBER> (<description>)"
OUTPUT_PATH = "/path/to/output/Test_Scenarios.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
CLR = {
    "navy":      "1F4E79",
    "navy_fg":   "FFFFFF",
    "blue_mid":  "2E75B6",
    "blue_lt":   "BDD7EE",
    "blue_pale": "DEEAF1",
    "nf":        "E2EFDA",
    "cb":        "D6DCE4",
    "neg":       "FCE4D6",
    "reg":       "FFF2CC",
    "p1":        "C00000",
    "p2":        "ED7D31",
    "p3":        "808080",
    "white":     "FFFFFF",
    "grey_lt":   "F2F2F2",
    "pass_bg":   "E2EFDA",
    "fail_bg":   "FCE4D6",
    "result_bg": "FFFBE6",
}

CAT_ROW_FILL = {
    "New Feature":        CLR["nf"],
    "Changed Behaviour":  CLR["cb"],
    "Negative":           CLR["neg"],
    "Regression":         CLR["reg"],
}

# ── Required keys for each scenario dict ─────────────────────────────────────
REQUIRED_SCENARIO_KEYS = {
    "id", "title", "priority", "category", "test_area", "type", "ac_coverage",
    "objective", "preconditions", "test_data", "pass_criteria", "fail_criteria",
}

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("TC ID",               12),
    ("Title",               42),
    ("Category",            16),
    ("Test Area",           20),
    ("Type",                16),
    ("Priority",             9),
    ("AC Coverage",         14),
    ("Objective",           42),
    ("Pre-conditions",      38),
    ("Test Data Required",  42),
    ("Pass Criteria",       36),
    ("Fail Criteria",       36),
    ("Notes / Gotchas",     36),
    ("Tester",              14),
    ("Exec Date",           13),
    ("Result",              16),
    ("Defect / Comment",    28),
]

# Column index lookup (1-based), derived from COLUMNS to stay in sync
COL = {label: i for i, (label, _) in enumerate(COLUMNS, 1)}

PRIORITY_COLORS = {
    "P1": CLR["p1"],
    "P2": CLR["p2"],
    "P3": CLR["p3"],
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _side(style="thin", color="000000"):
    return Side(style=style, color=color)


def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, italic=False, name="Calibri", underline=None):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name, underline=underline)


def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _estimate_height(text, col_width=38, line_height=14, min_height=28, max_height=180):
    """Estimate row height based on text content and column width."""
    if not text or not isinstance(text, str):
        return min_height
    lines = len(text) // col_width + text.count("\n") + 1
    return min(max(lines * line_height, min_height), max_height)


def validate_scenario(sc, index):
    """Validate that a scenario dict has all required keys."""
    missing = REQUIRED_SCENARIO_KEYS - set(sc.keys())
    if missing:
        raise ValueError(
            f"Scenario at index {index} (id={sc.get('id', '???')}) "
            f"is missing required keys: {sorted(missing)}"
        )


# ── Sheet construction helpers ───────────────────────────────────────────────
def _write_header(ws, ncols):
    """Write title and subtitle rows."""
    today = datetime.date.today().isoformat()

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=TITLE)
    t.fill = fill(CLR["navy"])
    t.font = _font(bold=True, color=CLR["navy_fg"], size=13)
    t.alignment = _align("center", v="center")
    t.border = thin_border()
    ws.row_dimensions[1].height = 26

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=f"{SUBTITLE}  |  Date: {today}")
    s.fill = fill(CLR["blue_lt"])
    s.font = _font(italic=True, color=CLR["navy"])
    s.alignment = _align("center", v="center")
    s.border = thin_border()
    ws.row_dimensions[2].height = 16


def _write_column_headers(ws):
    """Write column header row and set column widths."""
    for col, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=3, column=col, value=label)
        c.fill = fill(CLR["blue_mid"])
        c.font = _font(bold=True, color=CLR["navy_fg"])
        c.alignment = _align("center", v="center", wrap=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 22


def _write_data_row(ws, row, sc, row_index, dv):
    """Write a single scenario data row."""
    row_bg = CAT_ROW_FILL.get(sc["category"], CLR["white"])
    pri_bg = PRIORITY_COLORS.get(sc["priority"], CLR["p3"])

    values = [
        sc["id"], sc["title"], sc["category"], sc["test_area"],
        sc["type"], sc["priority"], sc["ac_coverage"],
        sc["objective"], sc["preconditions"], sc["test_data"],
        sc["pass_criteria"], sc["fail_criteria"], sc.get("notes", "—"),
        "", "", "", "",
    ]

    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin_border()
        c.alignment = _align()
        c.font = _font()
        c.fill = fill(row_bg)

        if col == COL["TC ID"]:
            c.fill = fill(CLR["blue_pale"])
            c.font = _font(bold=True, color=CLR["navy"], underline="single")
            c.alignment = _align("center", wrap=False)
        elif col == COL["Priority"]:
            c.fill = fill(pri_bg)
            c.font = _font(bold=True, color=CLR["navy_fg"])
            c.alignment = _align("center")
        elif col == COL["Pass Criteria"]:
            c.fill = fill(CLR["pass_bg"])
        elif col == COL["Fail Criteria"]:
            c.fill = fill(CLR["fail_bg"])
        elif col == COL["Exec Date"]:
            c.fill = fill(CLR["grey_lt"])
        elif col == COL["Result"]:
            c.fill = fill(CLR["result_bg"])
            dv.add(c)
        elif col == COL["Defect / Comment"]:
            c.fill = fill(CLR["white"])

    text_fields = [sc["preconditions"], sc["test_data"], sc.get("notes", "—"), sc["objective"]]
    ws.row_dimensions[row].height = max(
        _estimate_height(v) for v in text_fields if isinstance(v, str)
    )


def _write_legend(ws, scenario_count, ncols):
    """Write legend row below all data rows."""
    legend_row = 4 + scenario_count
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=ncols)
    lg = ws.cell(
        row=legend_row, column=1,
        value="LEGEND:   Green = New Feature   |   Blue-Grey = Changed Behaviour   "
              "|   Orange = Negative   |   Yellow = Regression   "
              "|   P1 = Must pass   |   P2 = Should pass   |   P3 = Nice to verify",
    )
    lg.fill = fill(CLR["grey_lt"])
    lg.font = _font(italic=True, size=9, color="595959")
    lg.alignment = _align(v="center")
    lg.border = thin_border()
    ws.row_dimensions[legend_row].height = 16


# ══════════════════════════════════════════════════════════════════════════════
#  SCENARIOS — populate one dict per test case
# ══════════════════════════════════════════════════════════════════════════════
SCENARIOS = [
    # {
    #     "id":            "TC-NF-01",
    #     "title":         "Short descriptive title",
    #     "priority":      "P1",
    #     "category":      "New Feature",
    #     "test_area":     "Staffing",
    #     "type":          "End-to-End",
    #     "ac_coverage":   "AC-1, AC-2",
    #     "objective":     "One sentence: what this test proves.",
    #     "preconditions": "1. First condition\n2. Second condition",
    #     "test_data":     "Data needed: ...\nSQL: SELECT ...",
    #     "pass_criteria": "Explicit pass statement.",
    #     "fail_criteria": "Explicit fail statement.",
    #     "notes":         "Caveats or gotchas.",
    # },
]

# ── Build workbook ────────────────────────────────────────────────────────────
def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Scenarios"
    ws.sheet_view.showGridLines = False

    ncols = len(COLUMNS)

    _write_header(ws, ncols)
    _write_column_headers(ws)

    # Result dropdown
    dv = DataValidation(
        type="list", formula1='"Pass,Fail,Blocked,Not Run,N/A"',
        allow_blank=True, showDropDown=False,
    )
    ws.add_data_validation(dv)

    # Data rows
    for i, sc in enumerate(SCENARIOS):
        _write_data_row(ws, 4 + i, sc, i, dv)

    _write_legend(ws, len(SCENARIOS), ncols)

    ws.freeze_panes = "A4"

    # Auto-filter on data columns (header row through last data row)
    last_data_row = 3 + len(SCENARIOS)
    if SCENARIOS:
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{last_data_row}"

    return wb


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not SCENARIOS:
        print("WARNING: SCENARIOS is empty — no data rows will be generated.")

    for i, sc in enumerate(SCENARIOS):
        validate_scenario(sc, i)

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir and not os.path.isdir(out_dir):
        print(f"ERROR: Output directory does not exist: {out_dir}", file=sys.stderr)
        sys.exit(1)

    wb = build()

    try:
        wb.save(OUTPUT_PATH)
    except OSError as e:
        print(f"ERROR: Failed to save workbook: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
